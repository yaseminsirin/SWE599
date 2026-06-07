"""Compare five hybrid-ranking weight configs; export Excel for professor report."""

from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand
from django.test import override_settings

from apps.search.services.ranking_evaluation import (
    EVALUATION_QUERIES,
    WEIGHT_CONFIGS,
    grade_relevance,
    mrr_at_k,
    ndcg_at_k,
    precision_at_k,
    rank_with_weights,
)
from apps.search.services.semantic_search import semantic_search_jobs


class Command(BaseCommand):
    help = (
        "Run professor-style hybrid ranking evaluation: 7 test queries, "
        "5 weight configs, NDCG/MRR/Precision metrics, Excel output."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default="evaluation_results/ranking_weight_comparison.xlsx",
            help="Path to .xlsx output (default: evaluation_results/ranking_weight_comparison.xlsx)",
        )
        parser.add_argument(
            "--top-k",
            type=int,
            default=10,
            help="Rank depth for metrics (default 10).",
        )
        parser.add_argument(
            "--pool-k",
            type=int,
            default=30,
            help="Candidate pool size from pgvector rerank (default 30).",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            self.stderr.write(
                self.style.ERROR(
                    "openpyxl is required. Install with: pip install openpyxl"
                )
            )
            raise SystemExit(1) from exc

        output_path = Path(options["output"])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        top_k = max(1, options["top_k"])
        pool_k = max(top_k, options["pool_k"])

        summary_rows: list[dict] = []
        detail_rows: list[dict] = []
        qrel_rows: list[dict] = []

        self.stdout.write("=== Hybrid ranking weight evaluation ===")
        self.stdout.write(f"Queries: {len(EVALUATION_QUERIES)} | Configs: {len(WEIGHT_CONFIGS)}")
        self.stdout.write("")

        with override_settings(SEMANTIC_SEARCH_CANDIDATE_POOL=pool_k):
            for query, query_note in EVALUATION_QUERIES:
                self.stdout.write(f"Query: {query!r}")
                try:
                    pool = semantic_search_jobs(
                        query,
                        top_k=pool_k,
                        tech_only=False,
                        return_reranked_pool=True,
                    )
                except Exception as exc:
                    self.stderr.write(self.style.ERROR(f"  search failed: {exc}"))
                    continue

                if not pool:
                    self.stdout.write("  (empty pool)")
                    continue

                graded: list[tuple[dict, int, str]] = []
                for item in pool:
                    grade = grade_relevance(query, item["job"])
                    graded.append((item, grade.score, grade.reason))
                    qrel_rows.append(
                        {
                            "query": query,
                            "job_id": item["job"].id,
                            "title": (item["job"].title or "")[:120],
                            "relevance_grade": grade.score,
                            "relevance_reason": grade.reason,
                            "semantic_score": round(float(item["semantic_score"]), 4),
                            "lexical_score": round(float(item.get("lexical_score", 0.0)), 4),
                            "role_alignment_score": round(
                                float(item.get("role_alignment_score", 0.0)), 4
                            ),
                        }
                    )

                for config_id, label, ws, wl, wr in WEIGHT_CONFIGS:
                    ranked = rank_with_weights(pool, ws=ws, wl=wl, wr=wr)
                    grades = [
                        grade_relevance(query, row["job"]).score for row in ranked[:top_k]
                    ]
                    summary_rows.append(
                        {
                            "query": query,
                            "config_id": config_id,
                            "config_label": label,
                            "semantic_weight": ws,
                            "lexical_weight": wl,
                            "role_weight": wr,
                            "pool_size": len(pool),
                            f"ndcg@{top_k}": round(ndcg_at_k(grades, top_k), 4),
                            f"mrr@{top_k}": round(mrr_at_k(grades, top_k), 4),
                            f"precision@{top_k}": round(
                                precision_at_k(grades, top_k), 4
                            ),
                            "highly_relevant_in_top3": sum(
                                1 for g in grades[:3] if g >= 3
                            ),
                        }
                    )

                    for rank_pos, row in enumerate(ranked[:top_k], start=1):
                        job = row["job"]
                        g = grade_relevance(query, job)
                        detail_rows.append(
                            {
                                "query": query,
                                "config_id": config_id,
                                "config_label": label,
                                "rank": rank_pos,
                                "job_id": job.id,
                                "title": (job.title or "")[:120],
                                "eval_score": round(float(row["eval_score"]), 4),
                                "semantic_score": round(float(row["semantic_score"]), 4),
                                "lexical_score": round(
                                    float(row.get("lexical_score", 0.0)), 4
                                ),
                                "role_alignment_score": round(
                                    float(row.get("role_alignment_score", 0.0)), 4
                                ),
                                "relevance_grade": g.score,
                                "relevance_reason": g.reason,
                            }
                        )

        # Aggregate per config
        aggregate_rows: list[dict] = []
        for config_id, label, ws, wl, wr in WEIGHT_CONFIGS:
            subset = [r for r in summary_rows if r["config_id"] == config_id]
            if not subset:
                continue
            n = len(subset)
            aggregate_rows.append(
                {
                    "config_id": config_id,
                    "config_label": label,
                    "semantic_weight": ws,
                    "lexical_weight": wl,
                    "role_weight": wr,
                    "queries_evaluated": n,
                    f"avg_ndcg@{top_k}": round(
                        sum(r[f"ndcg@{top_k}"] for r in subset) / n, 4
                    ),
                    f"avg_mrr@{top_k}": round(
                        sum(r[f"mrr@{top_k}"] for r in subset) / n, 4
                    ),
                    f"avg_precision@{top_k}": round(
                        sum(r[f"precision@{top_k}"] for r in subset) / n, 4
                    ),
                    "total_highly_relevant_top3": sum(
                        r["highly_relevant_in_top3"] for r in subset
                    ),
                }
            )

        aggregate_rows.sort(key=lambda r: r[f"avg_ndcg@{top_k}"], reverse=True)
        winner = aggregate_rows[0] if aggregate_rows else None

        wb = Workbook()
        bold = Font(bold=True)

        # Sheet 1: Methodology
        ws_method = wb.active
        ws_method.title = "Methodology"
        methodology = [
            ["JobSense AI — Hybrid Ranking Weight Evaluation"],
            [],
            ["Formula", "final_rank = w_sem*semantic + w_lex*lexical + w_role*role_alignment"],
            ["Semantic score", "Cosine similarity from pgvector (MiniLM embeddings)"],
            ["Lexical score", "Token overlap on title/body/category"],
            ["Role alignment", "Title phrase fit for stack+role queries"],
            [],
            ["Industry practice (OpenSearch, Azure AI Search, arXiv hybrid studies)"],
            ["1", "Build labeled query set + relevance grades (qrels 0–3)"],
            ["2", "Sweep weight combinations on same candidate pool"],
            ["3", "Measure NDCG@K, MRR@K, Precision@K"],
            ["4", "Pick config with best average NDCG on validation queries"],
            ["5", "Optional: Bayesian optimization per query (production scale)"],
            [],
            ["Why not 100% semantic?", "Vectors miss exact keywords (Python vs python)"],
            ["Why not 100% lexical?", "Misses synonyms (developer vs engineer)"],
            ["Why role weight?", "Disambiguates stack queries (backend vs staff product)"],
            [],
            ["Test dataset", f"{len(EVALUATION_QUERIES)} queries — see Test_Dataset sheet"],
            ["Configs compared", f"{len(WEIGHT_CONFIGS)} — see Weight_Configs sheet"],
            [],
            ["Best config (this run)", winner["config_label"] if winner else "n/a"],
            [
                f"Avg NDCG@{top_k}",
                winner[f"avg_ndcg@{top_k}"] if winner else "n/a",
            ],
        ]
        for row in methodology:
            ws_method.append(row)
        ws_method["A1"].font = bold

        # Sheet 2: Test dataset
        ws_data = wb.create_sheet("Test_Dataset")
        ws_data.append(["query", "query_type", "grading_scale"])
        for query, note in EVALUATION_QUERIES:
            ws_data.append([query, note, "0=irrelevant, 1=weak, 2=relevant, 3=highly relevant"])

        # Sheet 3: Weight configs
        ws_weights = wb.create_sheet("Weight_Configs")
        ws_weights.append(
            ["config_id", "label", "semantic_w", "lexical_w", "role_w", "rationale"]
        )
        rationales = {
            "W1": "Baseline: pure vector retrieval",
            "W2": "Baseline: pure keyword overlap",
            "W3": "Classic 50/50 hybrid (no role term)",
            "W4": "Literature neural-heavy (~60% semantic, OpenSearch studies)",
            "W5": "Current JobSense production default",
        }
        for config_id, label, wsem, wlex, wrole in WEIGHT_CONFIGS:
            ws_weights.append(
                [config_id, label, wsem, wlex, wrole, rationales.get(config_id, "")]
            )

        # Sheet 4: Summary aggregate (main result for professor)
        ws_agg = wb.create_sheet("Summary_Ranking")
        if aggregate_rows:
            headers = list(aggregate_rows[0].keys())
            ws_agg.append(headers)
            for row in aggregate_rows:
                ws_agg.append([row[h] for h in headers])
            for cell in ws_agg[1]:
                cell.font = bold

        # Sheet 5: Per-query metrics
        ws_sum = wb.create_sheet("Per_Query_Metrics")
        if summary_rows:
            headers = list(summary_rows[0].keys())
            ws_sum.append(headers)
            for row in summary_rows:
                ws_sum.append([row[h] for h in headers])

        # Sheet 6: Detailed rankings
        ws_det = wb.create_sheet("Detailed_Rankings")
        if detail_rows:
            headers = list(detail_rows[0].keys())
            ws_det.append(headers)
            for row in detail_rows:
                ws_det.append([row[h] for h in headers])

        # Sheet 7: Qrels
        ws_qrel = wb.create_sheet("Relevance_Labels")
        if qrel_rows:
            headers = list(qrel_rows[0].keys())
            ws_qrel.append(headers)
            for row in qrel_rows:
                ws_qrel.append([row[h] for h in headers])

        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"Wrote Excel report: {output_path}"))
        if winner:
            self.stdout.write(
                f"Best config: {winner['config_label']} "
                f"(avg NDCG@{top_k}={winner[f'avg_ndcg@{top_k}']})"
            )
